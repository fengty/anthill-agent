"""Document plugins — extract text from PDF, Word, and Excel.

The most common real-user workflow is "I have a document, summarise it"
or "extract these fields from this file." Reading text from common
office formats is the unblocker — without it, every doc workflow hits
a wall on step zero.

We stay within the same workspace sandbox as file_*: paths resolve
relative to ANTHILL_PLUGIN_WORKSPACE and cannot escape via .. or
absolute paths.

Optional deps: pypdf, python-docx, openpyxl. Install via:
    pip install 'anthill-agent[docs]'

Each plugin returns a clean message and ok=False if its library is
missing, rather than crashing the agent loop.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from anthill.plugins.base import Plugin, PluginResult
from anthill.plugins.filesystem import resolve_in_workspace as _resolve_safely


def _missing(library: str) -> PluginResult:
    return PluginResult(
        output=None,
        ok=False,
        error=(
            f"This plugin needs the [{library}] extra. "
            f"Install with: pip install 'anthill-agent[docs]'"
        ),
    )


class PdfReadPlugin(Plugin):
    name = "pdf_read"
    description = "Read text from a PDF file in the workspace."

    async def call(self, *, path: str, max_chars: int = 50_000, **_: Any) -> PluginResult:
        try:
            from pypdf import PdfReader
        except ImportError:
            return _missing("docs")

        try:
            abs_path = _resolve_safely(path)
        except PermissionError as e:
            return PluginResult(output=None, ok=False, error=str(e))
        if not abs_path.exists():
            return PluginResult(output=None, ok=False, error=f"no such file: {path}")
        if not abs_path.is_file():
            return PluginResult(output=None, ok=False, error=f"not a file: {path}")

        try:
            reader = PdfReader(str(abs_path))
            pages = []
            for page in reader.pages:
                pages.append(page.extract_text() or "")
            text = "\n".join(pages).strip()
        except Exception as e:  # noqa: BLE001
            return PluginResult(output=None, ok=False, error=f"pdf parse error: {e}")

        truncated = len(text) > max_chars
        return PluginResult(
            output=text[:max_chars],
            metadata={
                "path": str(abs_path),
                "pages": len(reader.pages),
                "char_count": len(text),
                "truncated": truncated,
            },
        )


class DocxReadPlugin(Plugin):
    name = "docx_read"
    description = "Read text from a .docx (Word) file in the workspace."

    async def call(self, *, path: str, max_chars: int = 50_000, **_: Any) -> PluginResult:
        try:
            from docx import Document
        except ImportError:
            return _missing("docs")

        try:
            abs_path = _resolve_safely(path)
        except PermissionError as e:
            return PluginResult(output=None, ok=False, error=str(e))
        if not abs_path.exists():
            return PluginResult(output=None, ok=False, error=f"no such file: {path}")

        try:
            doc = Document(str(abs_path))
            paragraphs = [p.text for p in doc.paragraphs if p.text]
            text = "\n".join(paragraphs).strip()
        except Exception as e:  # noqa: BLE001
            return PluginResult(output=None, ok=False, error=f"docx parse error: {e}")

        truncated = len(text) > max_chars
        return PluginResult(
            output=text[:max_chars],
            metadata={
                "path": str(abs_path),
                "paragraphs": len(paragraphs),
                "char_count": len(text),
                "truncated": truncated,
            },
        )


class XlsxReadPlugin(Plugin):
    name = "xlsx_read"
    description = "Read cells from an .xlsx (Excel) file in the workspace."

    async def call(
        self,
        *,
        path: str,
        sheet: str | None = None,
        max_rows: int = 200,
        **_: Any,
    ) -> PluginResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return _missing("docs")

        try:
            abs_path = _resolve_safely(path)
        except PermissionError as e:
            return PluginResult(output=None, ok=False, error=str(e))
        if not abs_path.exists():
            return PluginResult(output=None, ok=False, error=f"no such file: {path}")

        try:
            wb = load_workbook(str(abs_path), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            return PluginResult(output=None, ok=False, error=f"xlsx parse error: {e}")

        sheets = wb.sheetnames
        if sheet is not None:
            if sheet not in sheets:
                return PluginResult(
                    output=None,
                    ok=False,
                    error=f"sheet {sheet!r} not found. Available: {sheets}",
                )
            ws = wb[sheet]
        else:
            ws = wb.active

        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))

        wb.close()

        return PluginResult(
            output=rows,
            metadata={
                "path": str(abs_path),
                "sheet": ws.title,
                "sheets": sheets,
                "rows": len(rows),
                "truncated": ws.max_row > max_rows if ws.max_row else False,
            },
        )
